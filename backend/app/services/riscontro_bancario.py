"""
Riscontro Bancario — GIAMS
Analisi e confronto tra estratto conto bancario (Excel) e costi registrati in piattaforma.
"""
import re
from datetime import date, timedelta
from difflib import SequenceMatcher
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Parsing file Excel banca
# ---------------------------------------------------------------------------

# Mapping colonne atteso (case-insensitive, parziale)
_COL_ALIASES = {
    "data": ["data", "data operazione", "data op"],
    "operazione": ["operazione", "tipo operazione", "tipo"],
    "dettagli": ["dettagli", "descrizione", "causale", "descrizione operazione"],
    "conto": ["conto", "conto o carta", "conto/carta"],
    "contabilizzazione": ["contabilizzazione", "data contabile", "data valuta", "valuta"],
    "categoria": ["categoria"],
    "valuta": ["valuta", "divisa", "currency"],
    "importo": ["importo", "amount", "dare/avere", "movimento"],
}


def _match_header(header_text: str) -> str | None:
    """Trova il campo logico corrispondente a un'intestazione Excel."""
    h = (header_text or "").strip().lower()
    if not h:
        return None
    for campo, aliases in _COL_ALIASES.items():
        for alias in aliases:
            if alias == h or h.startswith(alias):
                return campo
    return None


def _parse_importo(val) -> float | None:
    """Converte un valore importo dal formato italiano/bancario a float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("\u20ac", "").replace("EUR", "").strip()
    if not s or s == "-":
        return None
    # Formato italiano: 1.234,56 -> 1234.56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_data(val) -> date | None:
    """Converte un valore data dal formato Excel/stringa a date."""
    if val is None:
        return None
    if hasattr(val, "date"):
        return val.date() if callable(getattr(val, "date")) else val.date
    if isinstance(val, date):
        return val
    s = str(val).strip()
    # Prova formati comuni
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_estratto_conto(file_bytes: bytes) -> list[dict]:
    """
    Legge un file Excel bancario e restituisce una lista di transazioni.
    Ogni transazione: {data, operazione, dettagli, conto, contabilizzazione, categoria, valuta, importo}
    """
    import io
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Trova la riga header (prima riga con almeno 3 match)
    header_idx = None
    col_map = {}
    for i, row in enumerate(rows):
        mapping = {}
        for j, cell in enumerate(row):
            campo = _match_header(str(cell) if cell else "")
            if campo and campo not in mapping:
                mapping[campo] = j
        if len(mapping) >= 3 and "importo" in mapping:
            header_idx = i
            col_map = mapping
            break

    if header_idx is None:
        raise ValueError(
            "Intestazioni non riconosciute nel file Excel. "
            "Verifica che il file contenga colonne come: Data, Operazione, Dettagli, Importo."
        )

    transazioni = []
    for row in rows[header_idx + 1:]:
        # Salta righe completamente vuote
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        importo_raw = row[col_map["importo"]] if col_map.get("importo") is not None and col_map["importo"] < len(row) else None
        importo = _parse_importo(importo_raw)
        if importo is None:
            continue  # Salta righe senza importo valido

        data_col = col_map.get("data")
        data_val = _parse_data(row[data_col] if data_col is not None and data_col < len(row) else None)

        cont_col = col_map.get("contabilizzazione")
        cont_val = _parse_data(row[cont_col] if cont_col is not None and cont_col < len(row) else None)

        def _get_str(campo):
            idx = col_map.get(campo)
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ""

        transazioni.append({
            "data": data_val,
            "operazione": _get_str("operazione"),
            "dettagli": _get_str("dettagli"),
            "conto": _get_str("conto"),
            "contabilizzazione": cont_val,
            "categoria": _get_str("categoria"),
            "valuta": _get_str("valuta"),
            "importo": importo,
        })

    wb.close()
    return transazioni


# ---------------------------------------------------------------------------
# Algoritmo di matching
# ---------------------------------------------------------------------------

def _normalizza_testo(s: str) -> str:
    """Normalizza testo per confronto fuzzy."""
    s = (s or "").lower().strip()
    s = re.sub(r"[^a-z0-9\s]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _similarita_testo(a: str, b: str) -> float:
    """Calcola similarita' tra due stringhe (0-1)."""
    na, nb = _normalizza_testo(a), _normalizza_testo(b)
    if not na or not nb:
        return 0.0
    return SequenceMatcher(None, na, nb).ratio()


def _match_importo(importo_banca: float, importo_costo: float, tolleranza: float = 0.02) -> bool:
    """
    Verifica se gli importi corrispondono.
    La banca usa importi negativi per le uscite, i costi sono positivi (salvo note credito).
    """
    # L'importo banca e' negativo per uscite -> confronta con valore assoluto
    banca_abs = abs(importo_banca)
    costo_abs = abs(importo_costo)
    return abs(banca_abs - costo_abs) <= tolleranza


def _match_data(data_banca: date, data_costo: date, giorni_tolleranza: int = 5) -> bool:
    """Verifica se le date corrispondono entro un margine di tolleranza."""
    if not data_banca or not data_costo:
        return False
    return abs((data_banca - data_costo).days) <= giorni_tolleranza


def esegui_riscontro(transazioni_banca: list[dict], costi_db: list[dict]) -> dict:
    """
    Esegue il riscontro bancario confrontando transazioni bancarie con costi registrati.

    Restituisce:
    - abbinati: lista di {banca, costo, score, note}
    - solo_banca: transazioni presenti solo in banca (non trovate nei costi)
    - solo_piattaforma: costi presenti solo in piattaforma (non trovati in banca)
    - statistiche: riepilogo numerico
    """
    # Filtra solo uscite dalla banca (importi negativi)
    uscite_banca = [t for t in transazioni_banca if t["importo"] is not None and t["importo"] < 0]

    # Indici disponibili (non ancora abbinati)
    banca_disponibili = set(range(len(uscite_banca)))
    costi_disponibili = set(range(len(costi_db)))

    abbinati = []

    # PASS 1: Match importo + data (alta confidenza)
    for bi in list(banca_disponibili):
        tb = uscite_banca[bi]
        best_ci = None
        best_score = 0
        best_delta_giorni = 999
        best_fornitore_sim = 0.0

        for ci in costi_disponibili:
            tc = costi_db[ci]
            # Confronta importo (banca negativo vs costo positivo/negativo)
            if not _match_importo(tb["importo"], tc["importo_totale"]):
                continue

            # Confronta data: banca vs data_pagamento del costo
            data_banca = tb.get("contabilizzazione") or tb.get("data")
            data_costo = tc.get("data_pagamento")

            if not _match_data(data_banca, data_costo, giorni_tolleranza=3):
                continue

            # Distanza data in giorni
            delta_giorni = abs((data_banca - data_costo).days) if data_banca and data_costo else 999

            # Similarita' testo tra operazione banca e fornitore costo
            fornitore_sim = max(
                _similarita_testo(tb.get("dettagli", ""), tc.get("fornitore", "")),
                _similarita_testo(tb.get("operazione", ""), tc.get("fornitore", "")),
            )
            desc_sim = max(
                fornitore_sim,
                _similarita_testo(tb.get("dettagli", ""), tc.get("descrizione", "")),
            )

            # Score basato su similarita' testo
            score = 80  # Base: importo + data matchano
            score += desc_sim * 20  # Bonus per similarita' descrizione

            if score > best_score:
                best_score = score
                best_ci = ci
                best_delta_giorni = delta_giorni
                best_fornitore_sim = fornitore_sim

        if best_ci is not None and best_score >= 75:
            # Verificato se:
            # 1) data esatta (stesso giorno) + importo uguale, OPPURE
            # 2) data vicina (max 2 gg) + importo uguale + fornitore trovato in descrizione banca
            data_esatta = best_delta_giorni == 0
            data_vicina = best_delta_giorni <= 2
            fornitore_match = best_fornitore_sim >= 0.4
            tipo = "verificato" if (data_esatta or (data_vicina and fornitore_match)) else "da_verificare"
            abbinati.append({
                "banca": uscite_banca[bi],
                "costo": costi_db[best_ci],
                "score": round(best_score),
                "tipo": tipo,
            })
            banca_disponibili.discard(bi)
            costi_disponibili.discard(best_ci)

    # PASS 2: Match solo importo con tolleranza data piu' ampia
    for bi in list(banca_disponibili):
        tb = uscite_banca[bi]
        best_ci = None
        best_score = 0

        for ci in costi_disponibili:
            tc = costi_db[ci]
            if not _match_importo(tb["importo"], tc["importo_totale"]):
                continue

            data_banca = tb.get("contabilizzazione") or tb.get("data")
            data_costo = tc.get("data_pagamento")

            score = 50  # Base: solo importo matcha
            if _match_data(data_banca, data_costo, giorni_tolleranza=10):
                score += 20

            desc_sim = max(
                _similarita_testo(tb.get("dettagli", ""), tc.get("descrizione", "")),
                _similarita_testo(tb.get("dettagli", ""), tc.get("fornitore", "")),
            )
            score += desc_sim * 15

            if score > best_score:
                best_score = score
                best_ci = ci

        if best_ci is not None and best_score >= 50:
            abbinati.append({
                "banca": uscite_banca[bi],
                "costo": costi_db[best_ci],
                "score": round(best_score),
                "tipo": "da_verificare",
            })
            banca_disponibili.discard(bi)
            costi_disponibili.discard(best_ci)

    # Risultati finali
    solo_banca = [uscite_banca[i] for i in sorted(banca_disponibili)]
    solo_piattaforma = [costi_db[i] for i in sorted(costi_disponibili)]

    totale_banca = sum(abs(t["importo"]) for t in uscite_banca)
    totale_piattaforma = sum(abs(c["importo_totale"]) for c in costi_db)
    totale_abbinati = sum(abs(m["banca"]["importo"]) for m in abbinati)
    totale_solo_banca = sum(abs(t["importo"]) for t in solo_banca)
    totale_solo_piattaforma = sum(abs(c["importo_totale"]) for c in solo_piattaforma)

    return {
        "abbinati": abbinati,
        "solo_banca": solo_banca,
        "solo_piattaforma": solo_piattaforma,
        "statistiche": {
            "transazioni_banca": len(uscite_banca),
            "costi_piattaforma": len(costi_db),
            "abbinati": len(abbinati),
            "solo_banca": len(solo_banca),
            "solo_piattaforma": len(solo_piattaforma),
            "totale_banca": round(totale_banca, 2),
            "totale_piattaforma": round(totale_piattaforma, 2),
            "totale_abbinati": round(totale_abbinati, 2),
            "totale_solo_banca": round(totale_solo_banca, 2),
            "totale_solo_piattaforma": round(totale_solo_piattaforma, 2),
            "differenza": round(totale_banca - totale_piattaforma, 2),
            "percentuale_copertura": round(len(abbinati) / len(uscite_banca) * 100, 1) if uscite_banca else 0,
        },
    }
